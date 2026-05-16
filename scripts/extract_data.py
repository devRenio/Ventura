"""Extract real data from source XLSX files into a JSON file we can hard-code.

Outputs: scripts/extracted_data.json with these keys:
  - population_by_dong:   [{dong, total, male, female, households, foreigners}, ...]
  - city_age_brackets:    [{label, total, male, female}, ...]          # 5-year cohorts
  - dong_age_breakdown:   {dong: [{label, total, male, female}, ...]}  # for top dong only
  - city_industry_top:    [{section, name, establishments, workers}, ...]
  - foreigner_age:        [{label, total, male, female}, ...]
"""
from openpyxl import load_workbook
from openpyxl.chartsheet import Chartsheet
from pathlib import Path
import json
import re

BASE = Path(
    r"c:\Users\eunho\Desktop\공은호\대학\3-1\비교과\협업하는프론티어\프로젝트 데이터"
)
OUT = Path(__file__).parent / "extracted_data.json"

result: dict = {}


def open_book(name: str):
    return load_workbook(BASE / name, data_only=True, read_only=False)


# ---------------------------------------------------------------
# 1) 주민등록인구 통계
# ---------------------------------------------------------------
wb_pop = open_book("2025. 12. 31.기준 주민등록인구통계표.xlsx")

# --- 표2: 동별 세대 및 인구 ---
ws = wb_pop["표2. 동별 세대 및 인구"]
rows = list(ws.iter_rows(values_only=True))
# headers begin at row index 4 (안양시 합계). dong rows have None in col A and dong name in col B.
population_by_dong = []
current_gu = None
for row in rows[4:]:
    a, b = row[0], row[1]
    if a in ("만안구", "동안구"):
        current_gu = a
        continue
    if a is None and isinstance(b, str) and b.strip():
        # b is dong name
        try:
            total = int(row[2] or 0)
            male = int(row[3] or 0)
            female = int(row[4] or 0)
            households = int(row[5] or 0)
            korean_total = int(row[6] or 0)
            foreigners = int(row[9] or 0)
        except (TypeError, ValueError):
            continue
        if total == 0:
            continue
        population_by_dong.append({
            "gu": current_gu,
            "dong": b.strip(),
            "total": total,
            "male": male,
            "female": female,
            "households": households,
            "koreans": korean_total,
            "foreigners": foreigners,
        })

result["population_by_dong"] = population_by_dong

# --- 표3: 안양시 전체 5세 단위 연령별 인구 ---
ws = wb_pop["표3. 연령별 인구(외국인제외)"]
rows = list(ws.iter_rows(values_only=True))
city_age_brackets = []
for row in rows[4:]:
    label = row[0]
    if not isinstance(label, str):
        continue
    label = label.strip()
    # keep 5-year brackets like "20~24세" and "85세 이상"
    if re.fullmatch(r"\d+~\d+세", label) or label.endswith("세 이상") or label == "계":
        try:
            total = int(row[1] or 0)
            male = int(row[2] or 0)
            female = int(row[3] or 0)
        except (TypeError, ValueError):
            continue
        city_age_brackets.append({"label": label, "total": total, "male": male, "female": female})

result["city_age_brackets"] = city_age_brackets

# --- 표4: 동별 연령별 인구 (외국인 제외) ---
ws = wb_pop["표4. 동별 연령별 인구(외국인제외)"]
rows = list(ws.iter_rows(values_only=True))
# Build column index map: scan row 2 (dong name) and row 3 (계/남/여)
header_dong = rows[2]
header_sex = rows[3]
dong_cols: dict[str, dict[str, int]] = {}
last_dong = None
for ci, val in enumerate(header_dong):
    if isinstance(val, str) and val.strip():
        last_dong = val.strip()
    sex = header_sex[ci] if ci < len(header_sex) else None
    if last_dong and isinstance(sex, str) and sex.strip() in ("계", "남", "여"):
        dong_cols.setdefault(last_dong, {})[sex.strip()] = ci

# pull 5-year cohorts per dong
dong_age_breakdown: dict[str, list] = {}
for dong_name, cols in dong_cols.items():
    if dong_name in ("안양시", "만안구", "동안구"):
        continue
    rows_out = []
    for row in rows[4:]:
        label = row[0]
        if not isinstance(label, str):
            continue
        label = label.strip()
        if not (re.fullmatch(r"\d+~\d+세", label) or label.endswith("세 이상")):
            continue
        try:
            total = int(row[cols.get("계", -1)] or 0)
            male = int(row[cols.get("남", -1)] or 0)
            female = int(row[cols.get("여", -1)] or 0)
        except (TypeError, ValueError, IndexError):
            continue
        rows_out.append({"label": label, "total": total, "male": male, "female": female})
    if rows_out:
        dong_age_breakdown[dong_name] = rows_out

result["dong_age_breakdown"] = dong_age_breakdown

# --- 표5: 연령별 외국인 ---
ws = wb_pop["표5. 연령별 외국인 현황"]
rows = list(ws.iter_rows(values_only=True))
foreigner_age = []
for row in rows[4:]:
    label = row[0]
    if not isinstance(label, str):
        continue
    label = label.strip()
    if re.fullmatch(r"\d+~\d+세", label) or label.endswith("세 이상"):
        try:
            total = int(row[1] or 0)
            male = int(row[2] or 0)
            female = int(row[3] or 0)
        except (TypeError, ValueError):
            continue
        foreigner_age.append({"label": label, "total": total, "male": male, "female": female})

result["foreigner_age"] = foreigner_age

wb_pop.close()


# ---------------------------------------------------------------
# 2) 산업세세분류 (대분류 / 중분류 일부만 추출)
# ---------------------------------------------------------------
wb_ind = open_book("1. 산업세세분류별 총괄(안양시).xlsx")
ws = wb_ind.active  # only one data sheet
rows = list(ws.iter_rows(values_only=True))

industry_sections = []  # 대분류 only
industry_mid = []       # 중분류 only (will pick top 15 by workers)
for row in rows[6:]:
    # cols: A=대 B=중 C=소 D=세 E=세세 F=name G=establishments H=workers
    code_l, code_m, code_s, code_d, code_dd, name = row[0], row[1], row[2], row[3], row[4], row[5]
    estab = row[6]
    workers = row[7]
    if not isinstance(name, str):
        continue
    # Strip leading spaces
    label = name.strip()
    # parse numbers (might be '…' or '-')
    def _num(v):
        if isinstance(v, (int, float)):
            return int(v)
        if isinstance(v, str):
            v = v.replace(",", "").strip()
            if v.isdigit():
                return int(v)
        return None
    e = _num(estab)
    w = _num(workers)

    # 대분류: 중분류 코드가 '**'인 행 (예: 'A.농업, ...(01~03)')
    if code_l and code_m == "**":
        industry_sections.append({
            "section": str(code_l),
            "name": label,
            "establishments": e,
            "workers": w,
        })
    # 중분류: 소분류가 '**'이고 대,중분류 코드가 모두 있는 행
    elif code_m and code_s == "**" and code_l and code_m != "**":
        industry_mid.append({
            "section": str(code_l),
            "mid": str(code_m),
            "name": label,
            "establishments": e,
            "workers": w,
        })

result["city_industry_sections"] = industry_sections
# pick top 15 mid categories by workers (drop None)
result["city_industry_top_mid"] = sorted(
    [r for r in industry_mid if r["workers"] is not None],
    key=lambda r: -(r["workers"] or 0),
)[:15]

wb_ind.close()


# ---------------------------------------------------------------
# 3) 안양시 사회조사 결과 통계표 (Chartsheet skip)
# ---------------------------------------------------------------
social: list = []
wb_soc = open_book("2025 안양시 사회조사 결과 통계표.xlsx")
for s in wb_soc.sheetnames:
    try:
        ws = wb_soc[s]
        if isinstance(ws, Chartsheet):
            continue
        rs = list(ws.iter_rows(values_only=True))
        social.append({"sheet": s, "rows": [list(r) for r in rs[:25]]})
    except Exception as e:
        social.append({"sheet": s, "error": str(e)})

# only dump sheet names + first row count to a side file so I can inspect
side = Path(__file__).parent / "social_dump.txt"
with side.open("w", encoding="utf-8") as f:
    for s in social:
        f.write("=" * 80 + "\n")
        f.write(s.get("sheet", "?") + "\n")
        f.write("=" * 80 + "\n")
        if "rows" in s:
            for i, r in enumerate(s["rows"]):
                trimmed = r[:10]
                f.write(f"r{i:02d}: {trimmed}\n")
        else:
            f.write("ERR: " + s.get("error", "") + "\n")
        f.write("\n")

wb_soc.close()


# Dump JSON
OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"Wrote {OUT}  ({OUT.stat().st_size:,} bytes)")
print(f"Dongs found: {len(result['population_by_dong'])}")
print(f"Industry sections: {len(result['city_industry_sections'])}")
print(f"Top mid-industries: {len(result['city_industry_top_mid'])}")
print(f"Social sheets dumped to: {side}")

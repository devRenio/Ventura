"""Quick inspection of source XLSX files for the Ventura demo.

Dumps sheet names, dimensions, and a head() preview to stdout
so we can decide which rows/columns to hard-code into dummyData.js.
"""
from openpyxl import load_workbook
from pathlib import Path

BASE = Path(
    r"c:\Users\eunho\Desktop\공은호\대학\3-1\비교과\협업하는프론티어\프로젝트 데이터"
)

FILES = [
    "1. 산업세세분류별 총괄(안양시).xlsx",
    "2025 안양시 사회조사 결과 통계표.xlsx",
    "2025. 12. 31.기준 주민등록인구통계표.xlsx",
]


def dump(file_name: str, head_rows: int = 20, max_cols: int = 12) -> None:
    path = BASE / file_name
    print("\n" + "=" * 90)
    print(f"FILE: {file_name}  ({path.stat().st_size:,} bytes)")
    print("=" * 90)
    wb = load_workbook(path, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # read_only sheets need iter_rows to compute size; cap it
        rows = list(ws.iter_rows(values_only=True))
        n_rows = len(rows)
        n_cols = max((len(r) for r in rows), default=0)
        print(f"\n--- Sheet: '{sheet_name}'  (rows={n_rows}, cols={n_cols}) ---")
        preview_rows = rows[:head_rows]
        for i, row in enumerate(preview_rows):
            trimmed = list(row)[:max_cols]
            print(f"r{i:02d}: {trimmed}")
    wb.close()


for f in FILES:
    try:
        dump(f)
    except Exception as e:  # noqa: BLE001
        print(f"!! Failed to read {f}: {e}")
